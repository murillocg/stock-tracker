"""Import a B3 *Negociação* export into the transaction ledger.

    .venv/bin/python scripts/import_b3.py ~/Downloads/negociacao-....xlsx --dry-run
    .venv/bin/python scripts/import_b3.py ~/Downloads/negociacao-....xlsx

The *Movimentação* export is the wrong file: it records custody movements and
leaves 111 settlement rows with no price at all. Negociação is the trade report.

Corporate actions are not in either export, so the adjustments below are entered
by hand. Each one was derived by reconciling the ledger against the holdings the
broker reports, and every one is justified in ADJUSTMENTS.
"""

import argparse
import datetime as dt
import hashlib
import re
import sys
from collections import defaultdict
from decimal import Decimal
from typing import Any

import boto3
import openpyxl

from shared.models import Currency, Transaction, TransactionType
from shared.positions import LedgerError, current_position
from shared.repository import DynamoDbTransactionRepository

SHEET = "Negociação"

FRACTIONAL = re.compile(r"^([A-Z]{4}\d{1,2})F$")
"""B3 books odd lots under a trailing F — BBAS3F is BBAS3. Same security, one
position; leaving the suffix on would split every holding in two."""

BROKERS = {"NU": "NU INVEST", "BTG": "BTG PACTUAL", "INTER": "INTER", "RICO": "RICO"}

ADJUSTMENTS: list[dict[str, Any]] = [
    # Corporate actions. Absent from both B3 exports, and each confirmed against
    # the holdings the broker reports.
    {
        "ticker": "BBAS3",
        "date": dt.date(2024, 4, 1),
        "type": TransactionType.BONUS,
        "quantity": "200",
        "broker": "NU INVEST",
        "note": "2:1 split. The price went 56.08 (26/03/2024) -> 27.62 (06/06/2024) with "
        "no trades in between, and doubling the 200 held then reproduces 600.",
    },
    {
        "ticker": "ITSA4",
        "date": dt.date(2026, 6, 30),
        "type": TransactionType.BONUS,
        "quantity": "8",
        "broker": "NU INVEST",
        "note": "Bonificação. Trades account for 1,092 of the 1,100 held.",
    },
    # Custody transfers, 30/07/2020, Inter -> BTG. Absent from the Negociação
    # export because they are not trades; taken from Movimentação. The IN price
    # is the average Inter held at that moment, so the cost travels with the
    # shares rather than being re-based.
    {
        "ticker": "BBAS3",
        "date": dt.date(2020, 7, 30),
        "type": TransactionType.TRANSFER_OUT,
        "quantity": "150",
        "broker": "INTER",
        "note": "Moved to BTG. Without this Inter still appears to hold them.",
    },
    {
        "ticker": "BBAS3",
        "date": dt.date(2020, 7, 30),
        "type": TransactionType.TRANSFER_IN,
        "quantity": "150",
        "unit_price": "29.09",
        "broker": "BTG PACTUAL",
        "note": "Arrived from Inter at Inter's average of 29.09 (100 @ 27.01 + 50 @ 33.25).",
    },
    {
        "ticker": "BPAC11",
        "date": dt.date(2020, 7, 30),
        "type": TransactionType.TRANSFER_OUT,
        "quantity": "80",
        "broker": "INTER",
        "note": "Moved to BTG.",
    },
    {
        "ticker": "BPAC11",
        "date": dt.date(2020, 7, 30),
        "type": TransactionType.TRANSFER_IN,
        "quantity": "80",
        "unit_price": "28.31",
        "broker": "BTG PACTUAL",
        "note": "Arrived from Inter at Inter's average of 28.31.",
    },
    {
        "ticker": "AXIA3",
        "date": dt.date(2026, 3, 1),
        "type": TransactionType.BUY,
        "quantity": "220",
        "unit_price": "40.54",
        "broker": "NU INVEST",
        "note": "Arrived through the Eletrobras restructuring, which appears in neither "
        "export. Quantity and average taken from the Nu Invest statement.",
    },
]
"""Everything else comes straight from the B3 file.

There were seven of these. Four were opening balances at a guessed price, needed
only so the fold could survive sales of shares bought before B3's data begins —
`current_position` treats a short position as a reset, which makes them
unnecessary. What is left are three real events, none of them invented.
"""

EXPECTED = {
    "AXIA3": 220,
    "BBAS3": 600,
    "BBDC3": 800,
    "BBSE3": 300,
    "BPAC11": 100,
    "CPLE3": 300,
    "GMAT3": 1000,
    "INBR32": 1000,
    "ITSA4": 1100,
    "MRVE3": 1200,
    "MSFT34": 60,
    "PRIO3": 600,
    "SPCX34": 50,
    "TSMC34": 58,
    "TTEN3": 800,
    "VALE3": 300,
    "VAMO3": 3000,
}
"""What the brokers say you hold. The import refuses to run if the ledger
disagrees — a silent mismatch is worse than no import."""


def normalise(code: object) -> str:
    match = FRACTIONAL.match(str(code).strip().upper())
    return match.group(1) if match else str(code).strip().upper()


def broker_of(name: object) -> str:
    upper = str(name).upper()
    if upper.startswith("NU"):
        return BROKERS["NU"]
    if "BTG" in upper:
        return BROKERS["BTG"]
    if "INTER" in upper:
        return BROKERS["INTER"]
    return BROKERS["RICO"]


def stable_id(ticker: str, date: dt.date, row: list[Any], index: int) -> str:
    """A repeatable id, so re-running an import overwrites instead of duplicating.

    The sort key is `<date>#<id>`. With the model's random default, every run
    would mint new ids and write a second copy of every row.
    """
    material = "|".join(str(x) for x in (ticker, date, row[1], row[6], row[7], row[4], index))
    return hashlib.sha1(material.encode()).hexdigest()[:12]


def read_trades(path: str) -> list[Transaction]:
    sheet = openpyxl.load_workbook(path)[SHEET]
    rows = (
        [sheet.cell(row=r, column=c).value for c in range(1, 10)]
        for r in range(2, sheet.max_row + 1)
    )
    trades = []
    for index, row in enumerate(rows):
        ticker = normalise(row[5])
        date = dt.datetime.strptime(str(row[0]), "%d/%m/%Y").date()
        trades.append(
            Transaction(
                ticker=ticker,
                date=date,
                type=TransactionType.BUY if row[1] == "Compra" else TransactionType.SELL,
                quantity=Decimal(str(row[6])),
                unit_price=Decimal(str(row[7])),
                currency=Currency.BRL,
                broker=broker_of(row[4]),
                # The file's own row order. B3 lists newest first, so this
                # descends through time — which is fine: it only ever breaks ties
                # within a single date, and it does so consistently.
                sequence=index,
                id=stable_id(ticker, date, row, index),
            )
        )
    return trades


def build_adjustments() -> list[Transaction]:
    return [
        Transaction(
            ticker=a["ticker"],
            date=a["date"],
            type=a["type"],
            quantity=Decimal(a["quantity"]),
            unit_price=Decimal(a.get("unit_price", "0")),
            currency=Currency.BRL,
            broker=a["broker"],
            note=a["note"],
            sequence=-1,  # adjustments settle before that day's trades
            id=hashlib.sha1(f"adj|{a['ticker']}|{a['date']}|{a['quantity']}".encode()).hexdigest()[
                :12
            ],
        )
        for a in ADJUSTMENTS
    ]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Import a B3 Negociação export.")
    parser.add_argument("path")
    parser.add_argument("--dry-run", action="store_true", help="Report without writing.")
    parser.add_argument("--table", default="stock-tracker-Transactions")
    parser.add_argument("--region", default="us-east-1")
    args = parser.parse_args(argv)

    transactions = read_trades(args.path) + build_adjustments()
    by_ticker: dict[str, list[Transaction]] = defaultdict(list)
    for t in transactions:
        by_ticker[t.ticker].append(t)

    print(
        f"{len(transactions)} rows across {len(by_ticker)} tickers "
        f"({len(ADJUSTMENTS)} of them hand-entered)\n"
    )

    print(f"{'ticker':<9}{'qty':>8}{'expected':>10}{'avg':>9}{'invested':>12}{'realised':>11}")
    print("-" * 60)
    mismatches = []
    for ticker in sorted(EXPECTED):
        try:
            position = current_position(ticker, by_ticker.get(ticker, []))
        except LedgerError as exc:
            mismatches.append(f"{ticker}: {exc}")
            print(f"{ticker:<9}  LEDGER ERROR")
            continue
        want = EXPECTED[ticker]
        got = int(position.quantity) if position else 0
        flag = "" if got == want else "   <-- MISMATCH"
        if got != want:
            mismatches.append(f"{ticker}: ledger {got}, broker says {want}")
        if position:
            print(
                f"{ticker:<9}{position.quantity:>8}{want:>10}{position.average_price or '—':>9}"
                f"{position.invested:>12}{position.realised_gain:>11}{flag}"
            )
        else:
            print(f"{ticker:<9}{0:>8}{want:>10}{flag}")

    if mismatches:
        print("\nRefusing to import — the ledger disagrees with your brokers:")
        for m in mismatches:
            print(f"  {m}")
        return 1

    print("\nEvery registered holding reconciles.")
    if args.dry_run:
        print("Dry run: nothing written.")
        return 0

    repository = DynamoDbTransactionRepository(
        boto3.resource("dynamodb", region_name=args.region).Table(args.table)
    )
    for transaction in transactions:
        repository.save(transaction)
    print(f"Wrote {len(transactions)} rows to {args.table}.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
